"""/admin/api/projects/{id}/postings + /admin/api/postings."""

from __future__ import annotations

import asyncio
import csv
import io
import re
import zipfile
import json
import uuid
from collections.abc import AsyncIterator

import structlog
from fastapi import (
    APIRouter,
    Depends,
    File,
    Form,
    HTTPException,
    Query,
    Request,
    UploadFile,
    status,
)
from fastapi.responses import StreamingResponse
from sqlalchemy import func, select, update
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from core.realtime import subscribe_run_events

from api.admin.middleware.auth import get_current_user
from api.admin.schemas.postings import (
    CreateLinkRunFileParams,
    CreateRunParams,
    parse_site_filter,
    CreateSpinRunParams,
    PostingRunResponse,
    QueueItem,
    QueueLinkCheckItem,
    QueueResponse,
    ResolveBulkRequest,
    RunProgressResponse,
    SpinOriginalRow,
    TextItemResponse,
    UpdateRunParams,
)
from core.config import settings
from core.db import get_db_read, get_db_write
from core.storage import storage
from api.common.pagination import DEFAULT_LIMIT, MAX_LIMIT, PaginatedResponse, encode_cursor
from domain.app_settings.service import get_app_settings
from domain.audit.service import record as audit_record
from domain.postings.service import (
    can_manage_run,
    can_view_run,
    count_active_runs_for_user,
    create_run,
    get_run,
    item_sort_key,
    list_runs_for_project,
    list_runs_for_viewer,
    list_text_items_for_run,
    request_cancel,
    request_pause,
    request_resume,
    retry_failed_items,
    run_progress_counts,
    soft_delete_run,
)
from infrastructure.db.models import (
    CELERY_PRIORITY_MAP,
    PostingRun,
    PostingRunStatus,
    RunTaskType,
    TextItem,
    TextItemStatus,
)
from domain.projects.service import can_manage_project, can_view_project, get_project
from infrastructure.db.models import AdminUser

log = structlog.get_logger(__name__)


def parse_tag_list(s: str | None) -> list[str]:
    """'News, sports ,EU' → ['News','sports','EU'] (через запятую/перенос, без
    lowercase — теги матчатся как есть). Дедуп с сохранением порядка."""
    if not s:
        return []
    seen: dict[str, None] = {}
    for raw in s.replace("\n", ",").split(","):
        t = raw.strip()
        if t:
            seen.setdefault(t, None)
    return list(seen)


def parse_domain_list(s: str | None) -> list[str]:
    """Свободный текст (через запятую/перенос/пробел) → нормализованные домены.
    Невалидные строки отбрасываются. Дедуп с сохранением порядка."""
    if not s:
        return []
    from domain.text_links.service import normalize_domain
    seen: dict[str, None] = {}
    for raw in s.replace("\n", ",").replace(" ", ",").split(","):
        d = normalize_domain(raw)
        if d:
            seen.setdefault(d, None)
    return list(seen)


async def _store_site_filter(session: AsyncSession, run_id: int,
                             site_langs: str | None, site_tlds: str | None,
                             site_tags: str | None = None,
                             site_domains: str | None = None,
                             site_domains_key: str | None = None) -> None:
    """Сохранить фильтр пула сайтов (lang/tld/tags/domains) в gen_params рана.
    No-op если всё пусто. site_domains_key — большой список файлом в MinIO."""
    langs = parse_site_filter(site_langs)
    tlds = parse_site_filter(site_tlds)
    tags = parse_tag_list(site_tags)
    domains = parse_domain_list(site_domains)
    if not langs and not tlds and not tags and not domains and not site_domains_key:
        return
    gp = dict((await session.scalar(
        select(PostingRun.gen_params).where(PostingRun.id == run_id))) or {})
    if langs:
        gp["site_langs"] = langs
    if tlds:
        gp["site_tlds"] = tlds
    if tags:
        gp["site_tags"] = tags
    if domains:
        gp["site_domains"] = domains
    elif site_domains_key:
        gp["site_domains_key"] = site_domains_key
    await session.execute(update(PostingRun).where(PostingRun.id == run_id).values(gen_params=gp))
    await session.commit()


# Один роутер с двумя prefix-ами: nested /projects/{id}/postings и flat /postings/{id}
project_postings_router = APIRouter(prefix="/projects", tags=["postings"])
postings_router = APIRouter(prefix="/postings", tags=["postings"])


@postings_router.post("/domain-list")
async def upload_domain_list_endpoint(
    file: UploadFile = File(..., description="txt/csv: домены пула (по одному в строке)"),
    viewer: AdminUser = Depends(get_current_user),
) -> dict:
    """Загрузить большой список доменов файлом для пула доступов прогона.
    Нормализуем + дедупим, кладём в MinIO, возвращаем ключ + count. Ключ
    передаётся в params рана как site_domains_key (вместо inline site_domains)."""
    raw = await file.read()
    if len(raw) > 20 * 1024 * 1024:
        raise HTTPException(status_code=413, detail="File too large (>20 MB)")
    domains = parse_domain_list(raw.decode("utf-8", errors="ignore"))
    if not domains:
        raise HTTPException(status_code=400, detail="Не нашёл валидных доменов в файле")
    key = f"domains-{uuid.uuid4().hex}/list.txt"
    storage.put_bytes(
        settings.MINIO_BUCKET_UPLOADS, key,
        "\n".join(domains).encode("utf-8"), content_type="text/plain",
    )
    log.info("postings.domain_list.uploaded", actor_id=viewer.id, count=len(domains), key=key)
    return {"key": key, "count": len(domains)}


# ─── List runs of project ────────────────────────────────────────────


def _apply_gen_progress(resp: PostingRunResponse, gen_params) -> PostingRunResponse:
    """Прокинуть прогресс генерации (gen_done/gen_total из gen_params) в ответ —
    питает красный бар генерации в UI/очереди (фаза UNPACKING)."""
    gp = gen_params or {}
    resp.gen_done = gp.get("gen_done")
    resp.gen_total = gp.get("gen_total")
    # Фильтр пула доступов — для инфо в UI
    resp.site_langs = gp.get("site_langs") or None
    resp.site_tlds = gp.get("site_tlds") or None
    resp.site_tags = gp.get("site_tags") or None
    _dom = gp.get("site_domains")
    resp.site_domains_count = len(_dom) if _dom else None
    resp.site_domains_file = bool(gp.get("site_domains_key"))
    return resp


@project_postings_router.get(
    "/{project_id}/postings",
    response_model=list[PostingRunResponse],
)
async def list_project_runs(
    project_id: int,
    viewer: AdminUser = Depends(get_current_user),
    session: AsyncSession = Depends(get_db_read),
) -> list[PostingRunResponse]:
    project = await get_project(session, project_id)
    if project is None:
        raise HTTPException(status_code=404, detail="Project not found")
    if not can_view_project(viewer, project):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Cannot view this project")

    rows = await list_runs_for_project(session, project_id=project_id, limit=200)
    return [_apply_gen_progress(PostingRunResponse.model_validate(r), r.gen_params) for r in rows]


# ─── Create run (multipart: file + JSON-params) ──────────────────────


def _resolve_publish_window(pub_from, pub_to, app_cfg):
    """Окно публикации прогона: per-run [pub_from, pub_to], если заданы ОБЕ
    даты; иначе — глобальный дефолт из app_settings (default_publish_from/to)."""
    if pub_from is not None and pub_to is not None:
        return pub_from, pub_to
    return app_cfg.default_publish_from, app_cfg.default_publish_to


@project_postings_router.post(
    "/{project_id}/postings",
    response_model=PostingRunResponse,
    status_code=status.HTTP_201_CREATED,
)
async def create_project_run(
    project_id: int,
    request: Request,
    file: UploadFile = File(..., description=".zip с .txt текстами"),
    params: str = Form(..., description="JSON: CreateRunParams"),
    viewer: AdminUser = Depends(get_current_user),
    session: AsyncSession = Depends(get_db_write),
) -> PostingRunResponse:
    # 1. Parse params JSON
    try:
        parsed = CreateRunParams.model_validate_json(params)
    except Exception as e:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"Invalid params: {e}") from e

    # 2. Project & access
    project = await get_project(session, project_id)
    if project is None:
        raise HTTPException(status_code=404, detail="Project not found")
    if not can_manage_project(viewer, project):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Cannot create runs in this project")

    # 3. Backpressure
    max_active = settings.MAX_ACTIVE_RUNS_PER_USER
    active_count = await count_active_runs_for_user(session, viewer.id)
    if active_count >= max_active:
        raise HTTPException(
            status_code=status.HTTP_429_TOO_MANY_REQUESTS,
            detail=f"Too many active runs ({active_count}). Limit per user: {max_active}",
        )

    # 4. File validation
    if not file.filename or not file.filename.lower().endswith(".zip"):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Only .zip files supported")

    contents = await file.read()
    max_size = 5 * 1024 * 1024 * 1024  # 5 GB
    if len(contents) > max_size:
        raise HTTPException(
            status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
            detail=f"Archive too large (>{max_size // (1024 * 1024)} MB)",
        )

    # 5. Reserve run_id by inserting placeholder, then upload + enqueue
    # Загружаем zip во временный bucket с предсказуемым ключом, привязанным к будущему run.
    # Используем uuid как идентификатор upload, run_id узнаем после insert.
    upload_uuid = uuid.uuid4().hex
    upload_key = f"upload-{upload_uuid}/source.zip"
    try:
        storage.put_bytes(
            settings.MINIO_BUCKET_UPLOADS,
            upload_key,
            contents,
            content_type="application/zip",
        )
    except Exception as e:
        log.exception("postings.upload_failed", error=str(e))
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to store archive: {e}",
        ) from e

    # 6. Подтягиваем системные дефолты concurrency/timeout/publish_window (super_admin задаёт)
    app_cfg = await get_app_settings(session)

    # 7. Создаём запись прогона со статусом unpacking
    pub_from, pub_to = _resolve_publish_window(parsed.publish_from, parsed.publish_to, app_cfg)
    run = await create_run(
        session,
        project=project,
        creator=viewer,
        name=parsed.name,
        publish_from=pub_from,
        publish_to=pub_to,
        concurrency=app_cfg.default_concurrency,
        timeout_seconds=app_cfg.default_timeout_seconds,
        priority=parsed.priority,
        scheduled_for=parsed.scheduled_for,
        spread_days=parsed.spread_days,
        source_archive_storage_key=upload_key,
        proxy_id=parsed.proxy_id,
        proxy_selector=parsed.proxy_selector,
        posting_method=parsed.posting_method,
        max_posts_per_site=parsed.max_posts_per_site,
        post_verify=parsed.post_verify,
    )
    await _store_site_filter(session, run.id, parsed.site_langs, parsed.site_tlds,
                             parsed.site_tags, parsed.site_domains, parsed.site_domains_key)

    log.info(
        "postings.created",
        run_id=run.id,
        project_id=project_id,
        actor_id=viewer.id,
        name=run.name,
        archive_size=len(contents),
    )
    await audit_record(
        session,
        actor=viewer,
        action="postings.create",
        resource_type="run",
        resource_id=run.id,
        request=request,
        changes={
            "project_id": project_id,
            "name": run.name,
            "priority": run.priority,
            "archive_size": len(contents),
            "proxy_id": run.proxy_id,
        },
    )

    # 8. Enqueue TaskIQ unpack task
    try:
        from workers.taskiq.unpack import unpack_archive

        await unpack_archive.kiq(run.id)
    except Exception as e:
        log.exception("postings.enqueue_failed", run_id=run.id, error=str(e))
        # Не падаем — статус остаётся unpacking, юзер увидит зависший run и сможет
        # передать ID админу для ручного восстановления. Recovery job из этапа 3
        # автоматически перезапустит зависшие unpacking-прогоны.

    return PostingRunResponse.model_validate(run)


@project_postings_router.post(
    "/{project_id}/postings/csv-direct",
    response_model=PostingRunResponse,
    status_code=status.HTTP_201_CREATED,
)
async def create_csv_direct_run(
    project_id: int,
    request: Request,
    file: UploadFile = File(..., description="csv/xlsx: link, anchor, text"),
    params: str = Form(..., description="JSON: CreateRunParams"),
    viewer: AdminUser = Depends(get_current_user),
    session: AsyncSession = Depends(get_db_write),
) -> PostingRunResponse:
    """Прямой вход: csv/xlsx со столбцами link, anchor, text → run с явными
    ссылками (без распаковки архива и без дизамбигуации)."""
    try:
        parsed = CreateRunParams.model_validate_json(params)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid params: {e}") from e
    project = await get_project(session, project_id)
    if project is None:
        raise HTTPException(status_code=404, detail="Project not found")
    if not can_manage_project(viewer, project):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Cannot create runs in this project")
    max_active = settings.MAX_ACTIVE_RUNS_PER_USER
    if await count_active_runs_for_user(session, viewer.id) >= max_active:
        raise HTTPException(status_code=429, detail=f"Too many active runs. Limit: {max_active}")
    fn = (file.filename or "").lower()
    if not fn.endswith((".csv", ".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="Only .csv or .xlsx supported")
    contents = await file.read()
    if len(contents) > 50 * 1024 * 1024:
        raise HTTPException(status_code=413, detail="File too large (>50 MB)")
    ext = "xlsx" if fn.endswith((".xlsx", ".xlsm")) else "csv"
    upload_key = f"upload-{uuid.uuid4().hex}/source.{ext}"
    try:
        storage.put_bytes(settings.MINIO_BUCKET_UPLOADS, upload_key, contents)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to store file: {e}") from e

    app_cfg = await get_app_settings(session)
    pub_from, pub_to = _resolve_publish_window(parsed.publish_from, parsed.publish_to, app_cfg)
    run = await create_run(
        session, project=project, creator=viewer, name=parsed.name,
        publish_from=pub_from, publish_to=pub_to,
        concurrency=app_cfg.default_concurrency, timeout_seconds=app_cfg.default_timeout_seconds,
        priority=parsed.priority, scheduled_for=parsed.scheduled_for,
        spread_days=parsed.spread_days, source_archive_storage_key=upload_key,
        proxy_id=parsed.proxy_id, proxy_selector=parsed.proxy_selector,
        posting_method=parsed.posting_method,
        max_posts_per_site=parsed.max_posts_per_site,
        post_verify=parsed.post_verify,
    )
    await _store_site_filter(session, run.id, parsed.site_langs, parsed.site_tlds,
                             parsed.site_tags, parsed.site_domains, parsed.site_domains_key)
    if parsed.csv_inject_link:
        # Флаг «инжектить ссылку из строки в текст» — читается воркером csv_direct.
        gp = (await session.scalar(
            select(PostingRun.gen_params).where(PostingRun.id == run.id))) or {}
        gp["csv_inject_link"] = True
        await session.execute(update(PostingRun).where(PostingRun.id == run.id)
                              .values(gen_params=gp))
        await session.commit()
    log.info("postings.created.csv_direct", run_id=run.id, project_id=project_id, actor_id=viewer.id)
    await audit_record(session, actor=viewer, action="postings.create_csv_direct",
                       resource_type="run", resource_id=run.id, request=request,
                       changes={"project_id": project_id, "name": run.name, "format": ext})
    try:
        from workers.taskiq.csv_direct import process_csv_direct
        await process_csv_direct.kiq(run.id)
    except Exception as e:
        log.exception("postings.csv_direct.enqueue_failed", run_id=run.id, error=str(e))
    return PostingRunResponse.model_validate(run)


@project_postings_router.post(
    "/{project_id}/postings/campaign",
    response_model=PostingRunResponse,
    status_code=status.HTTP_201_CREATED,
)
async def create_campaign_run(
    project_id: int,
    request: Request,
    file: UploadFile = File(..., description="csv/xlsx: anchor, link, count[, keyword, language]"),
    params: str = Form(..., description="JSON: CreateRunParams (с content_mode/run_mode)"),
    viewer: AdminUser = Depends(get_current_user),
    session: AsyncSession = Depends(get_db_write),
) -> PostingRunResponse:
    """csv_campaign (Content Engine): anchor,link,count[,keyword,language] →
    генерация (gen_per_post/gen_per_row) или reuse. Генерация идёт отдельной
    полосой (TaskIQ content.generate_campaign) и не блокирует постинг."""
    try:
        parsed = CreateRunParams.model_validate_json(params)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid params: {e}") from e
    if parsed.content_mode not in ("gen_per_post", "gen_per_row", "reuse"):
        raise HTTPException(status_code=400,
                            detail="content_mode обязателен: gen_per_post | gen_per_row | reuse")
    project = await get_project(session, project_id)
    if project is None:
        raise HTTPException(status_code=404, detail="Project not found")
    if not can_manage_project(viewer, project):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Cannot create runs in this project")
    if await count_active_runs_for_user(session, viewer.id) >= settings.MAX_ACTIVE_RUNS_PER_USER:
        raise HTTPException(status_code=429, detail="Too many active runs")
    fn = (file.filename or "").lower()
    if not fn.endswith((".csv", ".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="Нужен .csv или .xlsx (anchor,link,count)")
    contents = await file.read()
    if len(contents) > 50 * 1024 * 1024:
        raise HTTPException(status_code=413, detail="File too large (>50 MB)")

    from domain.content_engine import parse_content_csv
    pc = parse_content_csv(contents)
    if pc.error or not pc.rows:
        raise HTTPException(status_code=400,
                            detail=pc.error or "Нет валидных строк (нужно anchor,link,count)")

    upload_key = f"upload-{uuid.uuid4().hex}/campaign.csv"
    try:
        storage.put_bytes(settings.MINIO_BUCKET_UPLOADS, upload_key, contents)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to store file: {e}") from e

    app_cfg = await get_app_settings(session)
    pub_from, pub_to = _resolve_publish_window(parsed.publish_from, parsed.publish_to, app_cfg)
    run = await create_run(
        session, project=project, creator=viewer, name=parsed.name,
        publish_from=pub_from, publish_to=pub_to,
        concurrency=app_cfg.default_concurrency, timeout_seconds=app_cfg.default_timeout_seconds,
        priority=parsed.priority, scheduled_for=parsed.scheduled_for,
        spread_days=parsed.spread_days, source_archive_storage_key=upload_key,
        proxy_id=parsed.proxy_id, proxy_selector=parsed.proxy_selector,
        posting_method=parsed.posting_method,
        max_posts_per_site=parsed.max_posts_per_site,
        post_verify=parsed.post_verify,
    )
    # Оба режима (manual/auto) ПРЕД-СОЗДАЮТ пустые айтемы (видны в таблице) → READY,
    # Start ручной. Разница на Start: manual — отдельные «Сгенерировать»/«Старт
    # постинга»; auto — один Start стримит генерацию+постинг параллельно (Фаза 2).
    is_manual = parsed.run_mode == "manual"
    base_gp = {"rows": pc.rows, "prompt_template_id": parsed.prompt_template_id,
               "ai_model_id": parsed.ai_model_id, "language": parsed.language,
               **({"site_langs": parse_site_filter(parsed.site_langs)} if parse_site_filter(parsed.site_langs) else {}),
               **({"site_tlds": parse_site_filter(parsed.site_tlds)} if parse_site_filter(parsed.site_tlds) else {}),
               **({"site_tags": parse_tag_list(parsed.site_tags)} if parse_tag_list(parsed.site_tags) else {}),
               **({"site_domains": parse_domain_list(parsed.site_domains)} if parse_domain_list(parsed.site_domains)
                  else {"site_domains_key": parsed.site_domains_key} if parsed.site_domains_key else {})}
    from domain.content_engine import create_empty_campaign_items
    total, groups, main_ids = await create_empty_campaign_items(
        run.id, project.id, pc.rows, parsed.content_mode, parsed.language)
    gp = dict(base_gp)
    if groups:  # gen_per_row: спины заполнятся на Start/стриме
        gp.update({"fanout_groups": groups, "main_text_ids": main_ids, "deferred_fanout": True})
    # gen_total — число AI-генераций (оригиналы для gen_per_row, все для per_post)
    gp.update({"gen_done": 0, "gen_total": (len(main_ids) if groups else total)})
    await session.execute(update(PostingRun).where(PostingRun.id == run.id).values(
        content_source="csv_campaign", content_mode=parsed.content_mode,
        run_mode=parsed.run_mode, total_texts=total,
        status=PostingRunStatus.READY.value, gen_params=gp))
    await session.commit()
    log.info("postings.created.campaign", run_id=run.id, project_id=project_id,
             actor_id=viewer.id, mode=parsed.content_mode, items=total, manual=is_manual)
    await audit_record(session, actor=viewer, action="postings.create_campaign",
                       resource_type="run", resource_id=run.id, request=request,
                       changes={"project_id": project_id, "name": run.name,
                                "content_mode": parsed.content_mode, "run_mode": parsed.run_mode})
    fresh = await get_run(session, run.id)
    return _apply_gen_progress(PostingRunResponse.model_validate(fresh), fresh.gen_params)


# ─── Link runs (sitewide / homepage) ────────────────────────────────


@project_postings_router.get("/{project_id}/postings/link-candidates")
async def link_candidates_count(
    project_id: int,
    viewer: AdminUser = Depends(get_current_user),
    session: AsyncSession = Depends(get_db_read),
) -> dict:
    """Сколько сайтов получат сквозную ссылку (валидный administrator, нашей
    ссылки ещё нет)."""
    project = await get_project(session, project_id)
    if project is None:
        raise HTTPException(status_code=404, detail="Project not found")
    if not can_view_project(viewer, project):
        raise HTTPException(status_code=403, detail="No access")
    from domain.wp_links import count_candidate_link_sites

    return {"candidates": await count_candidate_link_sites(session)}


@project_postings_router.post(
    "/{project_id}/postings/links", status_code=status.HTTP_201_CREATED,
    response_model=PostingRunResponse,
)
async def create_link_run_endpoint(
    project_id: int,
    request: Request,
    file: UploadFile = File(..., description="csv/xlsx: anchor, link, count (count = на сколько сайтов)"),
    params: str = Form(..., description="JSON: {name, task_type, priority}"),
    viewer: AdminUser = Depends(get_current_user),
    session: AsyncSession = Depends(get_db_write),
) -> PostingRunResponse:
    """Создать run сквозной/homepage ссылки из файла anchor,link,count. count =
    на сколько сайтов поставить ссылку. Сайты — из пула администраторов (без
    пересечений), запускается потом обычным /start."""
    try:
        parsed = CreateLinkRunFileParams.model_validate_json(params)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid params: {e}") from e
    project = await get_project(session, project_id)
    if project is None:
        raise HTTPException(status_code=404, detail="Project not found")
    if not can_manage_project(viewer, project):
        raise HTTPException(status_code=403, detail="Cannot create runs in this project")
    fn = (file.filename or "").lower()
    if not fn.endswith((".csv", ".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="Нужен .csv или .xlsx (anchor,link,count)")
    contents = await file.read()
    if len(contents) > 50 * 1024 * 1024:
        raise HTTPException(status_code=413, detail="File too large (>50 MB)")

    from domain.content_engine import parse_content_csv
    pc = parse_content_csv(contents)
    if pc.error or not pc.rows:
        raise HTTPException(status_code=400,
                            detail=pc.error or "Нет валидных строк (нужно anchor,link,count)")
    links = [{"url": r["link"], "anchor": r.get("anchor") or "", "count": r.get("count") or 1}
             for r in pc.rows]

    app_cfg = await get_app_settings(session)
    pub_from, pub_to = _resolve_publish_window(parsed.publish_from, parsed.publish_to, app_cfg)
    from domain.wp_links import create_link_run

    run = await create_link_run(
        session, project=project, creator=viewer, name=parsed.name,
        task_type=parsed.task_type, links=links,
        concurrency=app_cfg.default_concurrency,
        timeout_seconds=app_cfg.default_timeout_seconds, priority=parsed.priority,
        site_langs=parse_site_filter(parsed.site_langs),
        site_tlds=parse_site_filter(parsed.site_tlds),
        site_tags=parse_tag_list(parsed.site_tags),
        site_domains=parse_domain_list(parsed.site_domains),
        site_domains_key=parsed.site_domains_key,
        max_posts_per_site=parsed.max_posts_per_site,
        proxy_selector=parsed.proxy_selector, spread_days=parsed.spread_days,
        scheduled_for=parsed.scheduled_for,
        publish_from=pub_from, publish_to=pub_to,
    )
    log.info("postings.link_run.created", run_id=run.id, project_id=project_id,
             task_type=parsed.task_type, links=len(links), total=run.total_texts, actor_id=viewer.id)
    await audit_record(
        session, actor=viewer, action="postings.create_link_run",
        resource_type="run", resource_id=run.id, request=request,
        changes={"project_id": project_id, "task_type": parsed.task_type,
                 "links": len(links), "total_items": run.total_texts},
    )
    fresh = await get_run(session, run.id)
    return PostingRunResponse.model_validate(fresh)


@project_postings_router.post(
    "/{project_id}/postings/spin", status_code=status.HTTP_201_CREATED,
    response_model=PostingRunResponse,
)
async def create_spin_run_endpoint(
    project_id: int,
    params: "CreateSpinRunParams",
    request: Request,
    viewer: AdminUser = Depends(get_current_user),
    session: AsyncSession = Depends(get_db_write),
) -> PostingRunResponse:
    """Создать spin_fanout-ран: M оригиналов-спинтаксов + размещения (link/anchor/
    count). manual → READY (ревью оригиналов → Start); auto → сразу fanout+постинг."""
    project = await get_project(session, project_id)
    if project is None:
        raise HTTPException(status_code=404, detail="Project not found")
    if not can_manage_project(viewer, project):
        raise HTTPException(status_code=403, detail="Cannot create runs in this project")
    from domain.content_engine import create_spin_run
    try:
        run = await create_spin_run(
            session, project_id=project_id, creator_id=viewer.id, name=params.name,
            originals=[o.model_dump() for o in params.originals],
            rows=[r.model_dump() for r in params.rows],
            run_mode=params.run_mode, scheduled_for=params.scheduled_for,
            spread_days=params.spread_days, proxy_selector=params.proxy_selector,
            posting_method=params.posting_method, priority=params.priority,
            site_langs=parse_site_filter(params.site_langs),
            site_tlds=parse_site_filter(params.site_tlds),
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e)) from e
    log.info("postings.spin_run.created", run_id=run.id, project_id=project_id,
             run_mode=params.run_mode, originals=len(params.originals), actor_id=viewer.id)
    await audit_record(session, actor=viewer, action="postings.create_spin_run",
                       resource_type="run", resource_id=run.id, request=request,
                       changes={"project_id": project_id, "run_mode": params.run_mode,
                                "originals": len(params.originals)})
    fresh = await get_run(session, run.id)
    return PostingRunResponse.model_validate(fresh)


@postings_router.get("/{run_id}/originals", response_model=list[SpinOriginalRow])
async def list_spin_originals(
    run_id: int,
    viewer: AdminUser = Depends(get_current_user),
    session: AsyncSession = Depends(get_db_read),
) -> list:
    """Оригиналы (спинтаксы) spin_fanout-рана для ревью перед Start."""
    run, _ = await _load_run_or_403(session, run_id, viewer, manage=False)
    from infrastructure.db.models import Text
    mids = (run.gen_params or {}).get("main_text_ids") or []
    if not mids:
        return []
    # карта оригинал → (link, anchor, count) из fanout_groups для показа в таблице
    groups = {g.get("text_id"): g for g in ((run.gen_params or {}).get("fanout_groups") or [])}
    rows = (await session.execute(
        select(Text).where(Text.id.in_(mids)).order_by(Text.id))).scalars().all()
    out = []
    for t in rows:
        g = groups.get(t.id) or {}
        out.append(SpinOriginalRow(
            id=t.id, title=t.title, lang=t.lang, spintax=t.body,
            link=g.get("link"), anchor=g.get("anchor") or None,
            placements=max(1, int(g.get("count") or 1))))
    return out


@postings_router.put("/{run_id}/originals/{text_id}", response_model=dict)
async def update_spin_original(
    run_id: int,
    text_id: int,
    payload: dict,
    viewer: AdminUser = Depends(get_current_user),
    session: AsyncSession = Depends(get_db_write),
) -> dict:
    """Правка ОРИГИНАЛА (до Start). payload: {body|spintax, title?}.

    Сохраняем как плайн-тело и СБРАСЫВАЕМ spin_formula → спинтакс пересоберётся
    из правленого тела на Start (deferred). Title не трогаем, если не передан.
    """
    run, _ = await _load_run_or_403(session, run_id, viewer, manage=True)
    mids = (run.gen_params or {}).get("main_text_ids") or []
    if text_id not in mids:
        raise HTTPException(status_code=404, detail="Original not in this run")
    if run.status not in (PostingRunStatus.READY.value, PostingRunStatus.DRAFT.value):
        raise HTTPException(status_code=409, detail="Можно править только до Start")
    new_body = (payload.get("body") or payload.get("spintax") or "").strip()
    if not new_body:
        raise HTTPException(status_code=400, detail="Пустое тело")
    from infrastructure.db.models import Text
    vals: dict = {"body": new_body, "spin_formula": None}  # спинтакс пересоберём на Start
    if "title" in payload:
        vals["title"] = (payload.get("title") or None)
    await session.execute(update(Text).where(Text.id == text_id).values(**vals))
    await session.commit()
    return {"ok": True, "text_id": text_id}


@postings_router.post("/{run_id}/text-items/{item_id}/remove-link",
                      status_code=status.HTTP_200_OK)
async def remove_link_endpoint(
    run_id: int,
    item_id: int,
    viewer: AdminUser = Depends(get_current_user),
    session: AsyncSession = Depends(get_db_write),
) -> dict:
    """Снять ранее размещённую сквозную ссылку (по placed_via+placement_ref)."""
    run, _ = await _load_run_or_403(session, run_id, viewer, manage=True)
    item = await session.scalar(
        select(TextItem).where(TextItem.id == item_id, TextItem.posting_run_id == run.id)
    )
    if item is None:
        raise HTTPException(status_code=404, detail="Item not found")
    from domain.wp_links import remove_link_item

    res = await remove_link_item(item_id, actor_id=viewer.id)
    await audit_record(
        session, actor=viewer, action="postings.remove_link",
        resource_type="text_item", resource_id=item_id,
        changes={"run_id": run_id, "status": res.get("status")},
    )
    return res


@postings_router.post("/{run_id}/resolve-bulk", status_code=status.HTTP_200_OK)
async def resolve_bulk_endpoint(
    run_id: int,
    payload: ResolveBulkRequest,
    viewer: AdminUser = Depends(get_current_user),
    session: AsyncSession = Depends(get_db_write),
) -> dict:
    """Массовый резолв needs_review-задач прогона по одному домену (каждой её
    собственная ссылка). Домен в проект НЕ добавляем — разовый резолв."""
    run, _ = await _load_run_or_403(session, run_id, viewer, manage=True)
    from domain.project_domains import resolve_run_by_domain

    try:
        res = await resolve_run_by_domain(session, run.id, payload.domain)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e)) from e
    await audit_record(
        session, actor=viewer, action="postings.resolve_bulk",
        resource_type="posting_run", resource_id=run.id,
        changes={"domain": payload.domain, **res},
    )
    return res


@postings_router.get("/{run_id}/needs-review-domains")
async def needs_review_domains_endpoint(
    run_id: int,
    viewer: AdminUser = Depends(get_current_user),
    session: AsyncSession = Depends(get_db_read),
) -> list[dict]:
    """Сводка доменов needs_review-задач прогона (для массового резолва на
    странице прогона). Пустой список = все нужные тексты уже привязаны."""
    run, _ = await _load_run_or_403(session, run_id, viewer)
    from domain.project_domains import needs_review_domains

    return await needs_review_domains(session, run.id)


@postings_router.post("/{run_id}/add-project-domain", status_code=status.HTTP_200_OK)
async def add_project_domain_for_run_endpoint(
    run_id: int,
    payload: ResolveBulkRequest,
    viewer: AdminUser = Depends(get_current_user),
    session: AsyncSession = Depends(get_db_write),
) -> dict:
    """Добавить домен в проект этого прогона → авто-резолв всех needs_review с ним
    (и будущих текстов). project_id берём с прогона — фронту передавать не нужно."""
    run, _ = await _load_run_or_403(session, run_id, viewer, manage=True)
    from domain.project_domains import add_domain

    try:
        res = await add_domain(session, run.project_id, payload.domain)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e)) from e
    await audit_record(
        session, actor=viewer, action="postings.add_project_domain",
        resource_type="posting_run", resource_id=run.id,
        changes={"domain": payload.domain, **res},
    )
    return res


async def _load_item_or_404(session: AsyncSession, run, item_id: int) -> TextItem:
    item = await session.scalar(
        select(TextItem).where(TextItem.id == item_id, TextItem.posting_run_id == run.id))
    if item is None:
        raise HTTPException(status_code=404, detail="Item not found")
    return item


@postings_router.post("/{run_id}/text-items/{item_id}/post",
                      status_code=status.HTTP_202_ACCEPTED)
async def post_item_endpoint(
    run_id: int,
    item_id: int,
    viewer: AdminUser = Depends(get_current_user),
    session: AsyncSession = Depends(get_db_write),
) -> dict:
    """Запостить ОДИН айтем по кнопке (юзер проверил текст и хочет запостить).
    Постит, перебирая сайты как общий цикл. Async — UI поллит статус айтема."""
    run, _ = await _load_run_or_403(session, run_id, viewer, manage=True)
    item = await _load_item_or_404(session, run, item_id)
    is_link = run.task_type in (RunTaskType.SITEWIDE_LINK.value,
                                RunTaskType.HOMEPAGE_LINK.value)
    if item.status == TextItemStatus.POSTED.value:
        raise HTTPException(status_code=409, detail="Уже запощен — используй Repost")
    if item.status not in (TextItemStatus.PENDING.value, TextItemStatus.FAILED.value):
        raise HTTPException(status_code=409,
                            detail=f"Нельзя постить в статусе '{item.status}'")
    if not is_link and item.text_id is None and item.storage_key is None:
        raise HTTPException(status_code=409, detail="У айтема нет текста — сначала сгенерируй")
    from core.celery_app import celery_app
    celery_app.send_task("postings.post_one_item", args=[item_id, False],
                         priority=CELERY_PRIORITY_MAP.get(run.priority, 5))
    await audit_record(session, actor=viewer, action="postings.post_item",
                       resource_type="text_item", resource_id=item_id,
                       changes={"run_id": run_id})
    return {"ok": True, "item_id": item_id, "status": "queued"}


@postings_router.post("/{run_id}/text-items/{item_id}/repost",
                      status_code=status.HTTP_202_ACCEPTED)
async def repost_item_endpoint(
    run_id: int,
    item_id: int,
    viewer: AdminUser = Depends(get_current_user),
    session: AsyncSession = Depends(get_db_write),
) -> dict:
    """Перезапостить уже запощенный айтем на ДРУГОЙ сайт (пост не отобразился на
    текущем доступе). Съедает ещё один слот сайта (max_posts_per_site)."""
    run, _ = await _load_run_or_403(session, run_id, viewer, manage=True)
    item = await _load_item_or_404(session, run, item_id)
    if item.status != TextItemStatus.POSTED.value:
        raise HTTPException(status_code=409, detail="Repost доступен только для запощенных")
    from core.celery_app import celery_app
    celery_app.send_task("postings.post_one_item", args=[item_id, True],
                         priority=CELERY_PRIORITY_MAP.get(run.priority, 5))
    await audit_record(session, actor=viewer, action="postings.repost_item",
                       resource_type="text_item", resource_id=item_id,
                       changes={"run_id": run_id, "prev_site_id": item.site_id})
    return {"ok": True, "item_id": item_id, "status": "queued"}


async def _enqueue_item_gen(session, run, item, *, regenerate: bool, viewer) -> dict:
    """Общая обвязка generate/regenerate: гварды + enqueue TaskIQ. Async — UI
    поллит статус (GENERATING → PENDING)."""
    if run.content_source != "csv_campaign":
        raise HTTPException(status_code=409, detail="Генерация только для AI-задач (csv_campaign)")
    if item.status in (TextItemStatus.POSTED.value, TextItemStatus.POSTING.value,
                       TextItemStatus.GENERATING.value):
        raise HTTPException(status_code=409, detail=f"Нельзя генерировать в статусе '{item.status}'")
    if regenerate and item.text_id is None:
        raise HTTPException(status_code=409, detail="Нет текста — используй «Сгенерировать»")
    if not regenerate and item.text_id is not None:
        raise HTTPException(status_code=409, detail="Текст уже есть — используй «Перегенерировать»")
    # Claim сразу → айтем мгновенно показывает «generating» (спиннер), а UI
    # поллит до возврата в стабильный статус. Защищает и от двойного клика (409).
    await session.execute(update(TextItem).where(TextItem.id == item.id)
                          .values(status=TextItemStatus.GENERATING.value, last_error=None))
    await session.commit()
    from workers.taskiq.cron_tasks import generate_item_task
    await generate_item_task.kiq(item.id, regenerate)
    await audit_record(session, actor=viewer,
                       action="postings.regenerate_item" if regenerate else "postings.generate_item",
                       resource_type="text_item", resource_id=item.id,
                       changes={"run_id": run.id})
    return {"ok": True, "item_id": item.id, "status": "queued"}


@postings_router.post("/{run_id}/text-items/{item_id}/generate",
                      status_code=status.HTTP_202_ACCEPTED)
async def generate_item_endpoint(
    run_id: int,
    item_id: int,
    viewer: AdminUser = Depends(get_current_user),
    session: AsyncSession = Depends(get_db_write),
) -> dict:
    """Сгенерировать текст для айтема без текста (AI / спин — зависит от режима)."""
    run, _ = await _load_run_or_403(session, run_id, viewer, manage=True)
    item = await _load_item_or_404(session, run, item_id)
    return await _enqueue_item_gen(session, run, item, regenerate=False, viewer=viewer)


@postings_router.post("/{run_id}/text-items/{item_id}/regenerate",
                      status_code=status.HTTP_202_ACCEPTED)
async def regenerate_item_endpoint(
    run_id: int,
    item_id: int,
    viewer: AdminUser = Depends(get_current_user),
    session: AsyncSession = Depends(get_db_write),
) -> dict:
    """Перегенерировать текст айтема (AI заново / переспин)."""
    run, _ = await _load_run_or_403(session, run_id, viewer, manage=True)
    item = await _load_item_or_404(session, run, item_id)
    return await _enqueue_item_gen(session, run, item, regenerate=True, viewer=viewer)


@postings_router.post("/{run_id}/generate-texts", status_code=status.HTTP_202_ACCEPTED)
async def generate_texts_endpoint(
    run_id: int,
    viewer: AdminUser = Depends(get_current_user),
    session: AsyncSession = Depends(get_db_write),
) -> dict:
    """Bulk «Сгенерировать тексты» для manual-задачи: фоном наполняет пустые
    предсозданные айтемы (gen_per_row → оригиналы, спины на Start; gen_per_post →
    все). Идемпотентно — генерит только то, что ещё пусто."""
    run, _ = await _load_run_or_403(session, run_id, viewer, manage=True)
    if run.content_source != "csv_campaign":
        raise HTTPException(status_code=409, detail="Только для AI-задач (csv_campaign)")
    # сколько ещё генерить: оригиналы (gen_per_row) либо все пустые (gen_per_post)
    groups = (run.gen_params or {}).get("fanout_groups") or []
    if groups:
        target_ids = [g["original_item_id"] for g in groups]
        empty = await session.scalar(select(func.count(TextItem.id)).where(
            TextItem.posting_run_id == run_id, TextItem.id.in_(target_ids),
            TextItem.text_id.is_(None)))
    else:
        empty = await session.scalar(select(func.count(TextItem.id)).where(
            TextItem.posting_run_id == run_id, TextItem.text_id.is_(None)))
    if not empty:
        raise HTTPException(status_code=409,
                            detail="Всё сгенерировано — точечно через пер-айтем перегенерацию")
    # UNPACKING = «активная обработка» (виден в очереди с оранжевым ген-баром).
    # gen_active=true ставим сразу (до старта таска), чтобы параллельный «Старт
    # постинга» гарантированно увидел идущую генерацию и не финишировал раньше.
    from domain.content_engine import set_gen_active
    await session.execute(update(PostingRun).where(PostingRun.id == run_id)
                          .values(status=PostingRunStatus.UNPACKING.value))
    await session.commit()
    await set_gen_active(run_id, True)
    from workers.taskiq.cron_tasks import generate_run_items_task
    await generate_run_items_task.kiq(run_id)
    await audit_record(session, actor=viewer, action="postings.generate_texts",
                       resource_type="run", resource_id=run_id)
    return {"ok": True, "run_id": run_id, "status": "generating"}


@postings_router.post("/{run_id}/fill-spins", status_code=status.HTTP_202_ACCEPTED)
async def fill_spins_endpoint(
    run_id: int,
    viewer: AdminUser = Depends(get_current_user),
    session: AsyncSession = Depends(get_db_write),
) -> dict:
    """Bulk «Заполнить спины» для manual gen_per_row: фоном расшивает готовые
    оригиналы в спин-варианты (видны в таблице для ревью) БЕЗ старта постинга.
    Идемпотентно — расшивает только то, что ещё не расшито."""
    run, _ = await _load_run_or_403(session, run_id, viewer, manage=True)
    if run.content_source != "csv_campaign" or run.content_mode != "gen_per_row":
        raise HTTPException(status_code=409,
                            detail="Только для AI-задач с генерацией на строку (спины)")
    if run.status not in (PostingRunStatus.READY.value, PostingRunStatus.DRAFT.value):
        raise HTTPException(status_code=409, detail="Задача не в статусе ready")
    # есть ли что расшивать: пустые спин-айтемы готовых оригиналов
    groups = (run.gen_params or {}).get("fanout_groups") or []
    ready_origs = set((await session.execute(select(TextItem.id).where(
        TextItem.id.in_([g["original_item_id"] for g in groups] or [0]),
        TextItem.text_id.isnot(None)))).scalars().all())
    spin_ids = [sid for g in groups if g.get("original_item_id") in ready_origs
                for sid in (g.get("spin_item_ids") or [])]
    empty = await session.scalar(select(func.count(TextItem.id)).where(
        TextItem.posting_run_id == run_id, TextItem.text_id.is_(None),
        TextItem.id.in_(spin_ids or [0]))) if spin_ids else 0
    if not empty:
        raise HTTPException(
            status_code=409,
            detail="Нет готовых оригиналов с пустыми спинами — сперва сгенерируйте тексты")
    # UNPACKING = «активная обработка» (оранжевый ген-бар, кнопки скрыты) — сразу,
    # чтобы UI не дал повторный клик до старта таска. Таск финиширует обратно в READY.
    await session.execute(update(PostingRun).where(PostingRun.id == run_id)
                          .values(status=PostingRunStatus.UNPACKING.value))
    await session.commit()
    from workers.taskiq.cron_tasks import fill_run_spins_task
    await fill_run_spins_task.kiq(run_id)
    await audit_record(session, actor=viewer, action="postings.fill_spins",
                       resource_type="run", resource_id=run_id)
    return {"ok": True, "run_id": run_id, "status": "filling"}


# ─── List all runs visible to viewer (cross-project dashboard) ──────


def _decode_cursor(cursor: str | None) -> int | None:
    if not cursor:
        return None
    import base64
    import json as _json

    try:
        raw = base64.urlsafe_b64decode(cursor.encode("ascii")).decode("utf-8")
        return int(_json.loads(raw)["after_id"])
    except Exception:
        return None


@postings_router.get("", response_model=PaginatedResponse[PostingRunResponse])
async def list_runs_endpoint(
    cursor: str | None = Query(default=None),
    limit: int = Query(default=DEFAULT_LIMIT, ge=1, le=MAX_LIMIT),
    statuses: list[str] = Query(default=[]),
    project_id: int | None = Query(default=None),
    created_by: int | None = Query(default=None),
    search: str | None = Query(default=None, max_length=200),
    viewer: AdminUser = Depends(get_current_user),
    session: AsyncSession = Depends(get_db_read),
) -> PaginatedResponse[PostingRunResponse]:
    after = _decode_cursor(cursor)
    rows = await list_runs_for_viewer(
        session,
        viewer=viewer,
        after_id=after,
        limit=limit,
        statuses=statuses or None,
        project_id=project_id,
        created_by=created_by,
        search=search,
    )
    has_more = len(rows) > limit
    if has_more:
        rows = rows[:limit]
    items = [_apply_gen_progress(PostingRunResponse.model_validate(r), r.gen_params) for r in rows]
    # cursor курсор — id последнего в обратном порядке (desc, поэтому after_id = последний.id)
    next_cursor = encode_cursor(rows[-1].id) if has_more and rows else None
    return PaginatedResponse[PostingRunResponse](items=items, next_cursor=next_cursor, has_more=has_more)


# ─── Global queue ────────────────────────────────────────────────────
# Видимая ВСЕМ юзерам очередь — игнорирует scope, чтобы команда могла
# координировать нагрузку (понимать сколько runs впереди их собственной).
# Только минимальная инфа: name + owner + project + status + progress.

# Global Queue = только то, что РЕАЛЬНО грузит сервер сейчас (оценить загрузку
# и сколько ещё разгребать). Спящие scheduled (drip между порциями) и paused
# (ручная пауза) НЕ показываем — иначе 100 фоновых drip-ранов забьют вывод.
_QUEUE_ACTIVE_STATUSES = (
    PostingRunStatus.UNPACKING.value,
    PostingRunStatus.QUEUED.value,
    PostingRunStatus.RUNNING.value,
)


@postings_router.get("/queue", response_model=QueueResponse)
async def global_queue_endpoint(
    viewer: AdminUser = Depends(get_current_user),
    session: AsyncSession = Depends(get_db_read),
) -> QueueResponse:
    """
    Глобальная очередь АКТИВНОЙ нагрузки по ВСЕМ юзерам/проектам — только то,
    что крутится/вот-вот закрутится (running/queued/unpacking). Спящие
    scheduled и paused сюда не попадают (см. _QUEUE_ACTIVE_STATUSES) — очередь
    отражает текущую загрузку сервера, а не весь бэклог.

    Сортировка:
      1. running first (то что прямо сейчас крутится)
      2. queued/unpacking — то что pickup-нется ближайшим воркером (FIFO по id)

    `is_mine` — true если creator_id == viewer.id, чтобы UI подсветил
    «свою» задачу.
    """
    from infrastructure.db.models import PostingRun, Project, AdminUser as AU

    # Левый join на project + creator чтобы отдавать сразу имена/usernames
    stmt = (
        select(
            PostingRun.id,
            PostingRun.name,
            PostingRun.status,
            PostingRun.priority,
            PostingRun.total_texts,
            PostingRun.posted_count,
            PostingRun.failed_count,
            PostingRun.gen_params,
            PostingRun.scheduled_for,
            PostingRun.started_at,
            PostingRun.created_at,
            PostingRun.created_by,
            Project.name.label("project_name"),
            AU.username.label("creator_username"),
        )
        .join(Project, Project.id == PostingRun.project_id)
        .outerjoin(AU, AU.id == PostingRun.created_by)
        .where(
            PostingRun.deleted_at.is_(None),
            PostingRun.status.in_(_QUEUE_ACTIVE_STATUSES),
        )
    )

    rows = list((await session.execute(stmt)).all())
    # Sort in Python — простой и предсказуемый порядок без сложного SQL:
    # running first, затем queued/unpacking (FIFO по id).
    status_order = {
        PostingRunStatus.RUNNING.value: 0,
        PostingRunStatus.QUEUED.value: 1,
        PostingRunStatus.UNPACKING.value: 1,
    }
    rows.sort(key=lambda r: (status_order.get(r.status, 99), r.id))

    items = [
        QueueItem(
            id=r.id,
            name=r.name,
            status=r.status,
            priority=r.priority,
            project_name=r.project_name,
            creator_username=r.creator_username,
            total_texts=r.total_texts or 0,
            posted_count=r.posted_count or 0,
            failed_count=r.failed_count or 0,
            gen_done=(r.gen_params or {}).get("gen_done"),
            gen_total=(r.gen_params or {}).get("gen_total"),
            scheduled_for=r.scheduled_for,
            started_at=r.started_at,
            created_at=r.created_at,
            is_mine=(r.created_by == viewer.id),
        )
        for r in rows
    ]

    # Активные перепроверки ссылок (link-check) — отдельная нагрузка на сервер,
    # чтобы очередь не выглядела пустой, пока идёт валидация завершённых прогонов.
    lc_rows = list((await session.execute(
        select(
            PostingRun.id, PostingRun.name, PostingRun.link_check_status,
            PostingRun.link_check_total, PostingRun.link_check_done,
            PostingRun.link_check_valid, PostingRun.created_by,
            Project.name.label("project_name"),
            AU.username.label("creator_username"),
        )
        .join(Project, Project.id == PostingRun.project_id)
        .outerjoin(AU, AU.id == PostingRun.created_by)
        .where(PostingRun.deleted_at.is_(None),
               PostingRun.link_check_status.in_(("queued", "running")))
        .order_by(PostingRun.id)
    )).all())
    link_checks = [
        QueueLinkCheckItem(
            id=r.id, name=r.name, project_name=r.project_name,
            creator_username=r.creator_username, status=r.link_check_status,
            total=r.link_check_total or 0, done=r.link_check_done or 0,
            valid=r.link_check_valid or 0, is_mine=(r.created_by == viewer.id),
        )
        for r in lc_rows
    ]
    return QueueResponse(items=items, total=len(items), link_checks=link_checks)


# ─── Get single run ──────────────────────────────────────────────────


@postings_router.get("/{run_id}", response_model=PostingRunResponse)
async def get_run_endpoint(
    run_id: int,
    viewer: AdminUser = Depends(get_current_user),
    session: AsyncSession = Depends(get_db_read),
) -> PostingRunResponse:
    run = await get_run(session, run_id)
    if run is None:
        raise HTTPException(status_code=404, detail="Run not found")
    project = await get_project(session, run.project_id)
    if project is None or not can_view_run(viewer, run, project):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Cannot view this run")
    resp = _apply_gen_progress(PostingRunResponse.model_validate(run), run.gen_params)
    # для csv_campaign — резолвим имена модели/промпта для шапки
    if run.content_source == "csv_campaign":
        from api.admin.schemas.postings import ContentParamsBrief
        from infrastructure.db.models import AiModel, PromptTemplate
        gp = run.gen_params or {}
        model_name = None
        if gp.get("ai_model_id"):
            model_name = await session.scalar(
                select(AiModel.display_name).where(AiModel.id == gp["ai_model_id"]))
        prompt_name = None
        if gp.get("prompt_template_id"):
            prompt_name = await session.scalar(
                select(PromptTemplate.name).where(PromptTemplate.id == gp["prompt_template_id"]))
        resp.content_params = ContentParamsBrief(
            language=gp.get("language"), model=model_name, prompt=prompt_name,
            error=gp.get("error"))
        # gen_per_row: сколько пустых спинов с готовым оригиналом можно расшить
        # (кнопка «Заполнить спины»). Только для manual-ready — кнопка там.
        if run.content_mode == "gen_per_row":
            groups = gp.get("fanout_groups") or []
            ready_origs = set((await session.execute(select(TextItem.id).where(
                TextItem.id.in_([g["original_item_id"] for g in groups] or [0]),
                TextItem.text_id.isnot(None)))).scalars().all())
            spin_ids = [sid for g in groups if g.get("original_item_id") in ready_origs
                        for sid in (g.get("spin_item_ids") or [])]
            resp.fillable_spins = (await session.scalar(select(func.count(TextItem.id)).where(
                TextItem.posting_run_id == run_id, TextItem.text_id.is_(None),
                TextItem.id.in_(spin_ids))) or 0) if spin_ids else 0
    return resp


# ─── Run detail: progress + text_items ──────────────────────────────


async def _load_run_or_403(session: AsyncSession, run_id: int, viewer: AdminUser, manage: bool = False):
    run = await get_run(session, run_id)
    if run is None:
        raise HTTPException(status_code=404, detail="Run not found")
    project = await get_project(session, run.project_id)
    if project is None:
        raise HTTPException(status_code=404, detail="Project not found")
    if manage:
        if not can_manage_run(viewer, run, project):
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN, detail="Cannot manage this run"
            )
    else:
        if not can_view_run(viewer, run, project):
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN, detail="Cannot view this run"
            )
    return run, project


@postings_router.get("/{run_id}/progress", response_model=RunProgressResponse)
async def get_run_progress(
    run_id: int,
    viewer: AdminUser = Depends(get_current_user),
    session: AsyncSession = Depends(get_db_read),
) -> RunProgressResponse:
    await _load_run_or_403(session, run_id, viewer)
    return RunProgressResponse(**await run_progress_counts(session, run_id))


# ─── SSE: live-прогресс run-а ──────────────────────────────────────


def _sse_format(event: str, data: str) -> bytes:
    """SSE protocol: event: name\\ndata: payload\\n\\n."""
    # data может быть multi-line — каждой строке свой `data:` префикс
    lines = [f"event: {event}"]
    for ln in data.splitlines() or [""]:
        lines.append(f"data: {ln}")
    lines.append("")  # пустая строка — конец события
    lines.append("")
    return ("\n".join(lines)).encode("utf-8")


@postings_router.get("/{run_id}/events")
async def stream_run_events(
    run_id: int,
    request: Request,
    viewer: AdminUser = Depends(get_current_user),
    session: AsyncSession = Depends(get_db_read),
) -> StreamingResponse:
    """
    Server-Sent Events: live progress + status changes для одного run-а.

    Поток событий:
    - `snapshot` сразу при коннекте — текущее состояние счётчиков и статуса.
    - `progress` после каждого text_item finalize (posted/failed/skipped).
    - `status` при смене статуса run-а (running/paused/done/...).
    - `ping` каждые ~20 сек чтобы proxy не закрыл idle-connection.

    Закрытие: клиент disconnect → asyncio.CancelledError → cleanup.
    """
    await _load_run_or_403(session, run_id, viewer)

    async def event_gen() -> AsyncIterator[bytes]:
        # 1. Snapshot
        progress = await run_progress_counts(session, run_id)
        run = await get_run(session, run_id)
        snapshot = {
            **progress,
            "status": run.status if run else None,
            "posted_count": run.posted_count if run else 0,
            "failed_count": run.failed_count if run else 0,
            "skipped_count": run.skipped_count if run else 0,
            "total_texts": run.total_texts if run else 0,
        }
        yield _sse_format("snapshot", json.dumps(snapshot, default=str))

        # 2. Live: race subscribe vs periodic ping vs client disconnect
        sub_iter = subscribe_run_events(run_id).__aiter__()
        sub_task: asyncio.Task | None = asyncio.create_task(sub_iter.__anext__())
        try:
            while True:
                if await request.is_disconnected():
                    break
                try:
                    done, _pending = await asyncio.wait(
                        {sub_task}, timeout=20.0, return_when=asyncio.FIRST_COMPLETED
                    )
                except asyncio.CancelledError:
                    break
                if not done:
                    # таймаут — keep-alive
                    yield b": ping\n\n"
                    continue
                try:
                    payload = sub_task.result()
                except StopAsyncIteration:
                    break
                except Exception as e:
                    log.warning("sse.subscribe.error", run_id=run_id, error=str(e))
                    break
                # payload — JSON-строка `{"event": "...", "data": {...}}`
                try:
                    parsed = json.loads(payload)
                    yield _sse_format(
                        parsed.get("event", "message"),
                        json.dumps(parsed.get("data", {}), default=str),
                    )
                except Exception:
                    yield _sse_format("message", payload)
                # Подготовить следующее ожидание
                sub_task = asyncio.create_task(sub_iter.__anext__())
        finally:
            if sub_task and not sub_task.done():
                sub_task.cancel()
            try:
                await sub_iter.aclose()  # type: ignore[attr-defined]
            except Exception:
                pass

    headers = {
        "Cache-Control": "no-cache",
        "X-Accel-Buffering": "no",  # nginx: не буферизовать
        "Connection": "keep-alive",
    }
    return StreamingResponse(event_gen(), media_type="text/event-stream", headers=headers)


@postings_router.get(
    "/{run_id}/text-items", response_model=PaginatedResponse[TextItemResponse]
)
async def list_run_text_items(
    run_id: int,
    cursor: str | None = Query(default=None),
    limit: int = Query(default=DEFAULT_LIMIT, ge=1, le=MAX_LIMIT),
    status_filter: str | None = Query(default=None, alias="status"),
    viewer: AdminUser = Depends(get_current_user),
    session: AsyncSession = Depends(get_db_read),
) -> PaginatedResponse[TextItemResponse]:
    await _load_run_or_403(session, run_id, viewer)
    after = _decode_cursor(cursor)
    rows = await list_text_items_for_run(
        session, run_id=run_id, status=status_filter, after_id=after, limit=limit
    )
    has_more = len(rows) > limit
    if has_more:
        rows = rows[:limit]
    items = [TextItemResponse.model_validate(r) for r in rows]
    # курсор = sort_key последнего (текст-вверху сортировка), не голый id
    next_cursor = encode_cursor(item_sort_key(rows[-1])) if has_more and rows else None
    return PaginatedResponse[TextItemResponse](
        items=items, next_cursor=next_cursor, has_more=has_more
    )


# ─── Run control actions ────────────────────────────────────────────


@postings_router.post("/{run_id}/start", status_code=status.HTTP_202_ACCEPTED)
async def start_run_endpoint(
    run_id: int,
    viewer: AdminUser = Depends(get_current_user),
    session: AsyncSession = Depends(get_db_write),
) -> dict:
    """
    Запустить run в работу. Допустимые исходные статусы:
      - READY (загружен и распакован, ждёт ручного Start)
      - SCHEDULED (юзер хочет запустить НЕМЕДЛЕННО, не дожидаясь scheduled_for)

    Действие: переводит status → QUEUED + отправляет Celery task.
    """
    run, _ = await _load_run_or_403(session, run_id, viewer, manage=True)
    allowed = {PostingRunStatus.READY.value, PostingRunStatus.SCHEDULED.value}
    # gen_per_post: «Старт постинга» можно нажать ПОВЕРХ идущей генерации
    # (UNPACKING) — постинг забирает готовые тексты, генерация наполняет остальные
    # параллельно. Айтемы gen_per_post самодостаточны (ссылка уже в теле), потому
    # их можно постить по мере готовности. gen_per_row требует fanout — не сюда.
    if run.content_source == "csv_campaign" and run.content_mode == "gen_per_post":
        allowed.add(PostingRunStatus.UNPACKING.value)
    if run.status not in allowed:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail=f"Cannot start run in status '{run.status}' (need {', '.join(sorted(allowed))})",
        )

    # spin_fanout: на старте сперва раскладываем оригиналы (spin+materialize),
    # потом обычная очередь. start_spin_run сам ставит QUEUED + Celery.
    if run.content_source == "spin_fanout":
        from domain.content_engine import start_spin_run
        res = await start_spin_run(run_id)
        log.info("postings.start_spin", run_id=run_id, actor_id=viewer.id, res=res)
        await audit_record(session, actor=viewer, action="postings.start",
                           resource_type="run", resource_id=run_id)
        return {"ok": bool(res.get("ok")), "run_id": run_id,
                "status": res.get("status", "queued")}

    # csv_campaign gen_per_row MANUAL: на Start расшиваем отревьюенные оригиналы
    # (спины из готовых оригиналов). AUTO стримит генерацию+постинг в run_posting
    # (генерация сама делает fanout), поэтому НЕ идёт сюда.
    if (run.content_source == "csv_campaign" and run.run_mode == "manual"
            and (run.gen_params or {}).get("deferred_fanout")):
        from domain.content_engine import start_campaign_fanout
        res = await start_campaign_fanout(run_id)
        log.info("postings.start_campaign_fanout", run_id=run_id, actor_id=viewer.id, res=res)
        await audit_record(session, actor=viewer, action="postings.start",
                           resource_type="run", resource_id=run_id)
        return {"ok": bool(res.get("ok")), "run_id": run_id,
                "status": res.get("status", "queued")}

    await session.execute(
        update(PostingRun)
        .where(PostingRun.id == run_id)
        .values(status=PostingRunStatus.QUEUED.value)
    )
    await session.commit()

    from core.celery_app import celery_app

    celery_app.send_task(
        "postings.run_posting",
        args=[run_id],
        priority=CELERY_PRIORITY_MAP.get(run.priority, 5),
    )
    log.info("postings.start_requested", run_id=run_id, actor_id=viewer.id)
    await audit_record(
        session, actor=viewer, action="postings.start",
        resource_type="run", resource_id=run_id,
    )
    return {"ok": True, "run_id": run_id, "status": PostingRunStatus.QUEUED.value}


@postings_router.post("/{run_id}/validate-links", status_code=status.HTTP_202_ACCEPTED)
async def validate_links_endpoint(
    run_id: int,
    viewer: AdminUser = Depends(get_current_user),
    session: AsyncSession = Depends(get_db_write),
) -> dict:
    """Перепроверить уже-валидные бэклинки завершённого прогона.

    Отдельная фоновая задача (TaskIQ) — перефетчит страницы постов и обновит
    отметку link_verified. Доступно ТОЛЬКО после завершения постинга (status=done).
    Видна в глобальной очереди как отдельный (фиолетовый) тип задач.
    """
    run, _ = await _load_run_or_403(session, run_id, viewer, manage=True)
    if run.status != PostingRunStatus.DONE.value:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="Link validation is available only after posting is finished (status 'done').",
        )
    if run.link_check_status == "running":
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="Link validation is already running for this run.",
        )
    target = int(await session.scalar(
        select(func.count(TextItem.id)).where(
            TextItem.posting_run_id == run_id,
            TextItem.status == TextItemStatus.POSTED.value,
            TextItem.link_verified.is_(True),
            TextItem.posted_url.isnot(None),
            TextItem.target_domain.isnot(None),
        )
    ) or 0)
    if target == 0:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="No verified links to re-validate in this run.",
        )
    await session.execute(
        update(PostingRun).where(PostingRun.id == run_id).values(
            link_check_status="queued",
            link_check_total=target,
            link_check_done=0,
            link_check_valid=0,
        )
    )
    await session.commit()

    from workers.taskiq.validate_links import validate_run_links
    await validate_run_links.kiq(run_id)
    log.info("postings.validate_links_requested", run_id=run_id, actor_id=viewer.id, total=target)
    await audit_record(
        session, actor=viewer, action="postings.validate_links",
        resource_type="run", resource_id=run_id,
    )
    return {"ok": True, "run_id": run_id, "total": target}


@postings_router.post("/{run_id}/restart", status_code=status.HTTP_202_ACCEPTED)
async def restart_run_endpoint(
    run_id: int,
    viewer: AdminUser = Depends(get_current_user),
    session: AsyncSession = Depends(get_db_write),
) -> dict:
    """
    Перезапустить упавший / прерванный / завершённый-с-failed run.

    Логика:
      1. Все text_items в статусах posting/failed/skipped → pending
         (worker подберёт заново; posted остаются posted)
      2. failed_count / skipped_count счётчики на run обнуляются
      3. Status → QUEUED + dispatch Celery task

    Доступно для статусов: FAILED, INTERRUPTED, CANCELLED, NEED_MORE_ADMINS,
    DONE (если есть failed items).
    """
    run, _ = await _load_run_or_403(session, run_id, viewer, manage=True)
    if run.status not in (
        PostingRunStatus.FAILED.value,
        PostingRunStatus.INTERRUPTED.value,
        PostingRunStatus.CANCELLED.value,
        PostingRunStatus.NEED_MORE_ADMINS.value,
        PostingRunStatus.DONE.value,
    ):
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail=f"Cannot restart run in status '{run.status}'",
        )

    from infrastructure.db.models import TextItem, TextItemStatus

    # Сбрасываем все НЕ-posted items в pending (включая зависшие в posting
    # и помеченные failed/skipped — даём им второй шанс)
    reset_res = await session.execute(
        update(TextItem)
        .where(
            TextItem.posting_run_id == run_id,
            TextItem.status.in_((
                TextItemStatus.POSTING.value,
                TextItemStatus.FAILED.value,
                TextItemStatus.SKIPPED.value,
            )),
        )
        .values(status=TextItemStatus.PENDING.value, last_error=None)
    )
    reset_count = int(reset_res.rowcount or 0)

    # Обнуляем счётчики неудачников на run-е (posted не трогаем)
    await session.execute(
        update(PostingRun)
        .where(PostingRun.id == run_id)
        .values(
            status=PostingRunStatus.QUEUED.value,
            failed_count=0,
            skipped_count=0,
            pause_requested=False,
            cancel_requested=False,
            finished_at=None,
        )
    )
    await session.commit()

    from core.celery_app import celery_app

    celery_app.send_task(
        "postings.run_posting",
        args=[run_id],
        priority=CELERY_PRIORITY_MAP.get(run.priority, 5),
    )
    log.info(
        "postings.restart", run_id=run_id, actor_id=viewer.id,
        items_reset=reset_count, prev_status=run.status,
    )
    await audit_record(
        session, actor=viewer, action="postings.restart",
        resource_type="run", resource_id=run_id,
    )
    return {
        "ok": True, "run_id": run_id,
        "status": PostingRunStatus.QUEUED.value,
        "items_reset": reset_count,
    }


def _set_or_del(gp: dict, key: str, value) -> None:
    """В gen_params: задать ключ при непустом значении, иначе удалить (clear)."""
    if value:
        gp[key] = value
    else:
        gp.pop(key, None)


@postings_router.patch("/{run_id}", response_model=PostingRunResponse)
async def update_run_endpoint(
    run_id: int,
    payload: UpdateRunParams,
    viewer: AdminUser = Depends(get_current_user),
    session: AsyncSession = Depends(get_db_write),
) -> PostingRunResponse:
    """Изменить настройки задачи. `max_posts_per_site` — в любом статусе (воркер
    читает live). Остальные постинг-параметры (приоритет, метод, окно публикации,
    расписание, drip, прокси, фильтр пула) — только пока задача ещё НЕ начала
    постить (READY / SCHEDULED). Меняются ТОЛЬКО явно переданные поля."""
    run, _ = await _load_run_or_403(session, run_id, viewer, manage=True)
    sent = payload.model_fields_set

    vals: dict = {}
    if "max_posts_per_site" in sent and payload.max_posts_per_site is not None:
        vals["max_posts_per_site"] = payload.max_posts_per_site

    deep_fields = {
        "priority", "scheduled_for", "spread_days", "posting_method", "post_verify",
        "proxy_selector", "publish_from", "publish_to",
        "site_langs", "site_tlds", "site_tags", "site_domains", "site_domains_key",
    }
    deep_sent = deep_fields & sent
    if deep_sent:
        if run.status not in (PostingRunStatus.READY.value, PostingRunStatus.SCHEDULED.value):
            raise HTTPException(
                status_code=409,
                detail="Параметры можно менять только до старта постинга (READY / SCHEDULED)",
            )
        for f in ("priority", "spread_days", "posting_method", "post_verify",
                  "proxy_selector", "publish_from", "publish_to"):
            if f in sent:
                vals[f] = getattr(payload, f)
        # scheduled_for + согласованный переход статуса
        if "scheduled_for" in sent:
            vals["scheduled_for"] = payload.scheduled_for
            if payload.scheduled_for is not None and run.status == PostingRunStatus.READY.value:
                vals["status"] = PostingRunStatus.SCHEDULED.value
            elif payload.scheduled_for is None and run.status == PostingRunStatus.SCHEDULED.value:
                vals["status"] = PostingRunStatus.READY.value
        # Фильтр пула → gen_params (мержим, контент-ключи кампании сохраняем)
        if {"site_langs", "site_tlds", "site_tags", "site_domains", "site_domains_key"} & sent:
            gp = dict(run.gen_params or {})
            if "site_langs" in sent:
                _set_or_del(gp, "site_langs", parse_site_filter(payload.site_langs))
            if "site_tlds" in sent:
                _set_or_del(gp, "site_tlds", parse_site_filter(payload.site_tlds))
            if "site_tags" in sent:
                _set_or_del(gp, "site_tags", parse_tag_list(payload.site_tags))
            if "site_domains" in sent:
                _set_or_del(gp, "site_domains", parse_domain_list(payload.site_domains))
            if "site_domains_key" in sent:
                _set_or_del(gp, "site_domains_key", payload.site_domains_key)
            vals["gen_params"] = gp

    if vals:
        await session.execute(
            update(PostingRun).where(PostingRun.id == run_id).values(**vals)
        )
        await session.commit()
        log.info("postings.updated", run_id=run_id, actor_id=viewer.id, fields=sorted(vals))
        await audit_record(session, actor=viewer, action="postings.update",
                           resource_type="run", resource_id=run_id,
                           changes={k: str(v)[:200] for k, v in vals.items()})
    fresh = await get_run(session, run_id)
    return _apply_gen_progress(PostingRunResponse.model_validate(fresh), fresh.gen_params)


@postings_router.post("/{run_id}/pause", status_code=status.HTTP_204_NO_CONTENT)
async def pause_run_endpoint(
    run_id: int,
    viewer: AdminUser = Depends(get_current_user),
    session: AsyncSession = Depends(get_db_write),
):
    run, _ = await _load_run_or_403(session, run_id, viewer, manage=True)
    if run.status not in (
        PostingRunStatus.RUNNING.value,
        PostingRunStatus.QUEUED.value,
    ):
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail=f"Cannot pause run in status '{run.status}'",
        )
    await request_pause(session, run_id)
    log.info("postings.pause_requested", run_id=run_id, actor_id=viewer.id)
    await audit_record(session, actor=viewer, action="postings.pause", resource_type="run", resource_id=run_id)


@postings_router.post("/{run_id}/resume", status_code=status.HTTP_204_NO_CONTENT)
async def resume_run_endpoint(
    run_id: int,
    viewer: AdminUser = Depends(get_current_user),
    session: AsyncSession = Depends(get_db_write),
):
    run, _ = await _load_run_or_403(session, run_id, viewer, manage=True)
    needs_enqueue = await request_resume(session, run_id)
    if needs_enqueue:
        from core.celery_app import celery_app

        celery_app.send_task(
            "postings.run_posting",
            args=[run_id],
            priority=CELERY_PRIORITY_MAP.get(run.priority, 5),
        )
    log.info("postings.resume_requested", run_id=run_id, actor_id=viewer.id)
    await audit_record(session, actor=viewer, action="postings.resume", resource_type="run", resource_id=run_id)


@postings_router.post("/{run_id}/cancel", status_code=status.HTTP_204_NO_CONTENT)
async def cancel_run_endpoint(
    run_id: int,
    viewer: AdminUser = Depends(get_current_user),
    session: AsyncSession = Depends(get_db_write),
):
    run, _ = await _load_run_or_403(session, run_id, viewer, manage=True)
    if run.status in (
        PostingRunStatus.DONE.value,
        PostingRunStatus.CANCELLED.value,
        PostingRunStatus.FAILED.value,
    ):
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail=f"Cannot cancel run in status '{run.status}'",
        )
    await request_cancel(session, run_id)
    log.info("postings.cancel_requested", run_id=run_id, actor_id=viewer.id)
    await audit_record(session, actor=viewer, action="postings.cancel", resource_type="run", resource_id=run_id)


@postings_router.post("/{run_id}/retry-failed")
async def retry_failed_endpoint(
    run_id: int,
    viewer: AdminUser = Depends(get_current_user),
    session: AsyncSession = Depends(get_db_write),
) -> dict:
    run, _ = await _load_run_or_403(session, run_id, viewer, manage=True)
    retried, needs_enqueue = await retry_failed_items(session, run_id)
    if needs_enqueue:
        from core.celery_app import celery_app

        celery_app.send_task(
            "postings.run_posting",
            args=[run_id],
            priority=CELERY_PRIORITY_MAP.get(run.priority, 5),
        )
    log.info("postings.retry_failed", run_id=run_id, actor_id=viewer.id, retried=retried)
    await audit_record(
        session, actor=viewer, action="postings.retry_failed",
        resource_type="run", resource_id=run_id, changes={"retried": retried},
    )
    return {"retried": retried, "re_enqueued": needs_enqueue}


@postings_router.delete("/{run_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_run_endpoint(
    run_id: int,
    viewer: AdminUser = Depends(get_current_user),
    session: AsyncSession = Depends(get_db_write),
):
    """
    Soft-delete: проставляет `deleted_at`. Run исчезает из списков, но БД-история
    остаётся (text_items, project_wp_used). Если run активен — сначала просим
    отмену (cancel_requested), иначе сразу soft-delete.
    """
    run, _ = await _load_run_or_403(session, run_id, viewer, manage=True)
    if run.status in (
        PostingRunStatus.RUNNING.value,
        PostingRunStatus.QUEUED.value,
        PostingRunStatus.UNPACKING.value,
        PostingRunStatus.PAUSED.value,
    ):
        # Сигнализируем воркеру остановиться, а сами архивируем — воркер
        # увидит deleted_at + cancel_requested и корректно завершит.
        await request_cancel(session, run_id)
    await soft_delete_run(session, run_id)
    log.info("postings.deleted", run_id=run_id, actor_id=viewer.id, prev_status=run.status)
    await audit_record(
        session, actor=viewer, action="postings.delete",
        resource_type="run", resource_id=run_id,
        changes={"prev_status": run.status},
    )


# ─── Экспорт результатов run-а ──────────────────────────────────────

_EXPORT_HEADER = [
    "text_id",
    "status",
    "title",
    "original_filename",
    "target_domain",
    "link_url",
    "link_anchor",
    "lang",
    "source",
    "site_domain",
    "credential_login",
    "posted_url",
    "link_verified",
    "post_id",
    "attempts",
    "posted_at",
    "last_error",
    "snippet",
]

_EXPORT_BATCH = 500  # сколько строк подгружаем за раз — не материализуем всё

_TAG_RE = re.compile(r"<[^>]+>")


def _snippet(body: str | None, n: int = 240) -> str:
    if not body:
        return ""
    return _TAG_RE.sub(" ", body).strip()[:n]


def _export_row(item: TextItem) -> list:
    """Возвращает row в порядке _EXPORT_HEADER. Используется CSV/XLSX/JSON."""
    txt = item.text if "text" in item.__dict__ else None
    return [
        item.id,
        item.status or "",
        (item.title or "").replace("\n", " ")[:500],
        item.original_filename or "",
        item.target_domain or "",
        item.link_url or "",
        (item.link_anchor or "").replace("\n", " ")[:500],
        item.lang or "",
        (txt.source if txt else "") or "",
        item.site.domain if item.site else "",
        item.credential.login if item.credential else "",
        item.posted_url or "",
        ("" if item.link_verified is None else ("valid" if item.link_verified else "invalid")),
        item.post_id,
        item.attempts or 0,
        item.posted_at.isoformat() if item.posted_at else "",
        (item.last_error or "").replace("\n", " ")[:1000],
        _snippet(txt.body if txt else None),
    ]


async def _iter_text_items(
    session: AsyncSession, run_id: int, *, verified_only: bool = False,
) -> AsyncIterator[TextItem]:
    after_id = 0
    while True:
        conds = [TextItem.posting_run_id == run_id, TextItem.id > after_id]
        if verified_only:
            # «Только валидные» — проставленные посты с подтверждённым бэклинком.
            conds += [
                TextItem.status == TextItemStatus.POSTED.value,
                TextItem.link_verified.is_(True),
            ]
        stmt = (
            select(TextItem)
            .where(*conds)
            .options(selectinload(TextItem.site), selectinload(TextItem.credential),
                     selectinload(TextItem.text))
            .order_by(TextItem.id)
            .limit(_EXPORT_BATCH)
        )
        rows = list((await session.execute(stmt)).scalars().all())
        if not rows:
            return
        for item in rows:
            yield item
        after_id = rows[-1].id
        if len(rows) < _EXPORT_BATCH:
            return


# ── CSV stream ────────────────────────────────────────────────────


async def _stream_run_csv(
    session: AsyncSession, run_id: int, *, verified_only: bool = False,
) -> AsyncIterator[bytes]:
    buf = io.StringIO()
    writer = csv.writer(buf, lineterminator="\n")
    writer.writerow(_EXPORT_HEADER)
    yield buf.getvalue().encode("utf-8")
    buf.seek(0)
    buf.truncate()

    n = 0
    async for item in _iter_text_items(session, run_id, verified_only=verified_only):
        writer.writerow([str(c) if c is not None else "" for c in _export_row(item)])
        n += 1
        if n % _EXPORT_BATCH == 0:
            chunk = buf.getvalue().encode("utf-8")
            buf.seek(0)
            buf.truncate()
            yield chunk
    if buf.tell() > 0:
        yield buf.getvalue().encode("utf-8")


# ── JSON stream ───────────────────────────────────────────────────


async def _stream_run_json(
    session: AsyncSession, run_id: int, *, verified_only: bool = False,
) -> AsyncIterator[bytes]:
    """JSON array, стримим построчно (`[{},\n{},\n…]`)."""
    yield b"[\n"
    first = True
    async for item in _iter_text_items(session, run_id, verified_only=verified_only):
        row = _export_row(item)
        obj = dict(zip(_EXPORT_HEADER, row, strict=True))
        prefix = b"" if first else b",\n"
        first = False
        yield prefix + json.dumps(obj, ensure_ascii=False).encode("utf-8")
    yield b"\n]\n"


# ── XLSX (binary, не стримим — собираем в BytesIO) ────────────────


async def _build_run_xlsx(
    session: AsyncSession, run_id: int, *, verified_only: bool = False,
) -> bytes:
    """
    openpyxl не поддерживает чистый stream-write красиво, поэтому собираем
    workbook в памяти. Для 100К строк это ~10 МБ, приемлемо. Если когда-то
    дорастём до миллионных runs — переедем на xlsxwriter с constant_memory=True.
    """
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = f"run-{run_id}"
    ws.append(_EXPORT_HEADER)
    async for item in _iter_text_items(session, run_id, verified_only=verified_only):
        ws.append(_export_row(item))

    # posted_url как hyperlink (нагляднее в Excel)
    url_col_idx = _EXPORT_HEADER.index("posted_url") + 1
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=url_col_idx)
        if cell.value:
            cell.hyperlink = cell.value
            cell.style = "Hyperlink"

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# ── Endpoint ──────────────────────────────────────────────────────


async def _build_run_txt_zip(
    session: AsyncSession, run_id: int, *, verified_only: bool = False,
) -> bytes:
    """Zip из .txt: 1 файл = постированное тело (texts.body материализованного
    варианта). Имя = {item_id}_{домен}.txt (item_id уникален → коллизий нет)."""
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        async for item in _iter_text_items(session, run_id, verified_only=verified_only):
            body = item.text.body if item.text else ""
            dom = item.target_domain or (item.site.domain if item.site else "")
            base = re.sub(r"[^A-Za-z0-9._-]+", "_", f"{item.id}_{dom}").strip("_") or str(item.id)
            zf.writestr(f"{base}.txt", body or "")
    return buf.getvalue()


@postings_router.get("/{run_id}/result")
async def download_run_result(
    run_id: int,
    format: str = Query(default="csv", pattern="^(csv|xlsx|json|txt)$"),
    verified_only: bool = Query(default=False),
    viewer: AdminUser = Depends(get_current_user),
    session: AsyncSession = Depends(get_db_read),
) -> StreamingResponse:
    await _load_run_or_403(session, run_id, viewer)

    # Суффикс имени файла, чтобы «только валидные» не перепутать с полной выгрузкой.
    suffix = "-valid" if verified_only else ""

    if format == "txt":
        data = await _build_run_txt_zip(session, run_id, verified_only=verified_only)
        return StreamingResponse(
            iter([data]), media_type="application/zip",
            headers={"Content-Disposition": f'attachment; filename="run-{run_id}{suffix}-texts.zip"'},
        )
    if format == "csv":
        return StreamingResponse(
            _stream_run_csv(session, run_id, verified_only=verified_only),
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": f'attachment; filename="run-{run_id}{suffix}.csv"'},
        )
    if format == "json":
        return StreamingResponse(
            _stream_run_json(session, run_id, verified_only=verified_only),
            media_type="application/json; charset=utf-8",
            headers={"Content-Disposition": f'attachment; filename="run-{run_id}{suffix}.json"'},
        )
    # xlsx
    data = await _build_run_xlsx(session, run_id, verified_only=verified_only)
    return StreamingResponse(
        iter([data]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="run-{run_id}{suffix}.xlsx"'},
    )
