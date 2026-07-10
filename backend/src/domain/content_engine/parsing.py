"""Парсер входных CSV для Content Engine + детект формата.

Два формата (по заголовкам):
  • direct   — готовые тексты: столбцы link, anchor, text
  • campaign — кампания (генерация/reuse): anchor, link, count[, keyword,
               language, content_parametrs]

Возвращаем структурированные строки; что с ними делать (готовый текст vs
генерация/spin-fanout) решает флоу создания рана.
"""

from __future__ import annotations

import csv
import io
import json
from dataclasses import dataclass, field


@dataclass
class ParsedCsv:
    fmt: str                       # 'direct' | 'campaign'
    rows: list[dict] = field(default_factory=list)
    skipped: int = 0
    error: str | None = None


# Алиасы заголовков → канон. Принимаем и единственное, и множественное число,
# и частые синонимы — чтобы файл «просто работал».
_HEADER_ALIASES = {
    "counts": "count", "qty": "count", "amount": "count",
    "links": "link", "url": "link", "urls": "link",
    "anchors": "anchor", "anchor_text": "anchor",
    "keywords": "keyword", "kw": "keyword",
    "lang": "language", "languages": "language",
    "html": "text", "snippet": "text", "body": "text",
}


def _norm_headers(fieldnames) -> dict[str, str]:
    """map нормализованный→оригинальный заголовок (lowercase/strip + алиасы)."""
    out: dict[str, str] = {}
    for f in fieldnames or []:
        key = (f or "").strip().lower()
        key = _HEADER_ALIASES.get(key, key)
        out[key] = f
    return out


def detect_format(headers: set[str]) -> str | None:
    if "text" in headers and "link" in headers:
        return "direct"
    if "count" in headers and "anchor" in headers and "link" in headers:
        return "campaign"
    return None


def _clean_cell(v) -> str:
    """Ячейка → строка. Целочисленные float (90.0) → '90', а не '90.0'."""
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _read_records(data: bytes) -> tuple[list[str], list[dict]]:
    """(fieldnames, строки-словари) из CSV или XLSX (определяем по magic-байтам)."""
    if data[:2] == b"PK":  # xlsx/xlsm = zip-контейнер
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        ws = wb.active
        it = ws.iter_rows(values_only=True)
        try:
            header_row = next(it)
        except StopIteration:
            return [], []
        fieldnames = [(_clean_cell(c)) for c in header_row]
        records: list[dict] = []
        for row in it:
            if row is None:
                continue
            rec = {fn: _clean_cell(row[i] if i < len(row) else None)
                   for i, fn in enumerate(fieldnames) if fn}
            if any(rec.values()):  # пропускаем полностью пустые строки
                records.append(rec)
        return fieldnames, records
    # CSV
    text = data.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    fieldnames = list(reader.fieldnames or [])
    records = [{k: (v or "") for k, v in r.items()} for r in reader]
    return fieldnames, records


def parse_content_csv(data: bytes) -> ParsedCsv:
    """Распарсить CSV/XLSX кампании или прямого входа в структурированные строки."""
    try:
        fieldnames, records = _read_records(data)
        hmap = _norm_headers(fieldnames)
        fmt = detect_format(set(hmap))
        if not fmt:
            return ParsedCsv(fmt="", error=(
                "Не распознан формат. Нужны столбцы либо link,anchor,text (готовые "
                "тексты), либо anchor,link,count[,keyword,language] (кампания)."))

        def g(row, key):
            col = hmap.get(key)
            return (row.get(col) or "").strip() if col else ""

        rows: list[dict] = []
        skipped = 0
        for raw in records:
            link = g(raw, "link")
            anchor = g(raw, "anchor")
            if fmt == "direct":
                body = g(raw, "text")
                if not link or not body:
                    skipped += 1
                    continue
                rows.append({"link": link, "anchor": anchor, "text": body})
            else:  # campaign
                if not link or not anchor:
                    skipped += 1
                    continue
                try:
                    count = max(1, int(float(g(raw, "count") or "1")))
                except ValueError:
                    count = 1
                cp_raw = g(raw, "content_parametrs") or g(raw, "content_params")
                cp = {}
                if cp_raw:
                    try:
                        cp = json.loads(cp_raw)
                    except Exception:
                        cp = {}
                rows.append({
                    "anchor": anchor, "link": link, "count": count,
                    "keyword": g(raw, "keyword"),
                    "language": g(raw, "language") or None,
                    "content_parametrs": cp,
                })
        return ParsedCsv(fmt=fmt, rows=rows, skipped=skipped)
    except Exception as e:
        return ParsedCsv(fmt="", error=f"Ошибка парсинга файла: {e}")


def parse_link_csv(data: bytes) -> ParsedCsv:
    """Парсер входа для link-ранов (сквозная/homepage).

    Столбцы: `anchor, link, count[, text]`.
      • text (он же html/snippet) — готовый HTML-сниппет со встроенной ссылкой:
        ставим КАК ЕСТЬ. Иначе строим <a href=link>anchor</a>.
      • строка валидна если есть `link` ИЛИ `text` (в text ссылка уже внутри).
      • count опц., дефолт 1 (= на сколько разных сайтов поставить).

    Возвращает rows: [{link, anchor, count, html}].
    """
    try:
        fieldnames, records = _read_records(data)
        hmap = _norm_headers(fieldnames)

        def g(row, key):
            col = hmap.get(key)
            return (row.get(col) or "").strip() if col else ""

        rows: list[dict] = []
        skipped = 0
        for raw in records:
            link = g(raw, "link")
            html = g(raw, "text")
            anchor = g(raw, "anchor")
            if not link and not html:
                skipped += 1
                continue
            try:
                count = max(1, int(float(g(raw, "count") or "1")))
            except ValueError:
                count = 1
            rows.append({"link": link, "anchor": anchor, "count": count, "html": html})
        if not rows:
            return ParsedCsv(fmt="link", skipped=skipped, error=(
                "Нет валидных строк. Нужен столбец link (или text — готовый HTML "
                "со встроенной ссылкой); count опционален."))
        return ParsedCsv(fmt="link", rows=rows, skipped=skipped)
    except Exception as e:
        return ParsedCsv(fmt="", error=f"Ошибка парсинга файла: {e}")
