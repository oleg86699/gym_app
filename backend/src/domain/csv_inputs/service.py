"""Парсеры табличных входов задач (B2).

csv-direct: csv/xlsx со столбцами link, anchor, text — данные заданы напрямую.
Заголовки регистронезависимы; лишние столбцы игнорируются. Строки без link или
text пропускаются.
"""

from __future__ import annotations

import csv
import io

import structlog

log = structlog.get_logger(__name__)

_REQUIRED = ("link", "anchor", "text")


def _norm_header(h: str | None) -> str:
    return (h or "").strip().lower()


def _rows_from_csv(content: bytes) -> list[dict]:
    text = content.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        return []
    header = [_norm_header(c) for c in rows[0]]
    idx = {col: header.index(col) for col in _REQUIRED if col in header}
    if "link" not in idx or "text" not in idx:
        raise ValueError("CSV must have columns: link, anchor, text")
    out: list[dict] = []
    for r in rows[1:]:
        def cell(col: str) -> str:
            i = idx.get(col)
            return (r[i].strip() if i is not None and i < len(r) else "")
        out.append({"link": cell("link"), "anchor": cell("anchor"), "text": cell("text")})
    return out


def _rows_from_xlsx(content: bytes) -> list[dict]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return []
    header = [_norm_header(str(c) if c is not None else "") for c in header_row]
    idx = {col: header.index(col) for col in _REQUIRED if col in header}
    if "link" not in idx or "text" not in idx:
        raise ValueError("XLSX must have columns: link, anchor, text")
    out: list[dict] = []
    for r in rows_iter:
        def cell(col: str) -> str:
            i = idx.get(col)
            v = r[i] if (i is not None and i < len(r)) else None
            return (str(v).strip() if v is not None else "")
        out.append({"link": cell("link"), "anchor": cell("anchor"), "text": cell("text")})
    wb.close()
    return out


def _rows_generic(content: bytes, filename: str) -> list[list[str]]:
    """Сырые строки (включая заголовок) из csv/xlsx как list[list[str]]."""
    fn = (filename or "").lower()
    if fn.endswith(".csv"):
        text = content.decode("utf-8-sig", errors="replace")
        return [[(c or "") for c in row] for row in csv.reader(io.StringIO(text))]
    if fn.endswith((".xlsx", ".xlsm")):
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        out = [[("" if c is None else str(c)) for c in row]
               for row in ws.iter_rows(values_only=True)]
        wb.close()
        return out
    raise ValueError("Only .csv or .xlsx supported")


def parse_campaign(content: bytes, filename: str) -> list[dict]:
    """Кампания: csv/xlsx со столбцами links|link, anchor, counts|count.
    Возвращает [{link, anchor, count}]. type (старый link_type) игнорируем.
    Строки без link пропускаем; count<1 → 1."""
    rows = _rows_generic(content, filename)
    if not rows:
        return []
    header = [_norm_header(c) for c in rows[0]]

    def col(*names: str) -> int | None:
        for n in names:
            if n in header:
                return header.index(n)
        return None

    li, ai, ci = col("links", "link"), col("anchor", "anchors"), col("counts", "count")
    if li is None:
        raise ValueError("Campaign file must have a 'links' (or 'link') column")
    out: list[dict] = []
    for r in rows[1:]:
        def cell(i: int | None) -> str:
            return (r[i].strip() if i is not None and i < len(r) else "")
        link = cell(li)
        if not link:
            continue
        cnt_raw = cell(ci)
        try:
            cnt = max(1, int(float(cnt_raw))) if cnt_raw else 1
        except ValueError:
            cnt = 1
        out.append({"link": link, "anchor": cell(ai), "count": cnt})
    log.info("csv_inputs.campaign_parsed", filename=filename, rows=len(out))
    return out


def parse_link_anchor_text(content: bytes, filename: str) -> list[dict]:
    """Распарсить csv/xlsx со столбцами link, anchor, text. Возвращает только
    валидные строки (есть link и text). Бросает ValueError если нет нужных
    столбцов / неподдерживаемый формат."""
    fn = (filename or "").lower()
    if fn.endswith(".csv"):
        rows = _rows_from_csv(content)
    elif fn.endswith((".xlsx", ".xlsm")):
        rows = _rows_from_xlsx(content)
    else:
        raise ValueError("Only .csv or .xlsx supported")
    valid = [r for r in rows if r.get("link") and r.get("text")]
    log.info("csv_inputs.parsed", filename=filename, total=len(rows), valid=len(valid))
    return valid
